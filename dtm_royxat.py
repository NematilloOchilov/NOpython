#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from time import sleep
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
for i in range(1, 16):
    print('--------', i)
    url = f"https://mandat.dtm.uz/Home/AfterFilter?page={str(i)}&"
    url = url + "region=14&university=324&faculty=5230100&edLang=1&edType=1&nog=False&muy=False&soldier=False&iiv=False&prez=0&notC=0&bans=0&covid=0&sortorder=ResultDesc"
    USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.14; rv:65.0) Gecko/20100101 Firefox/65.0"
    headers = {"user-agent": USER_AGENT}
    page = requests.get(url, headers=headers)
    table_rows = BeautifulSoup(page.text, 'html.parser').find_all('table', class_='table')[0].find_all('tr')
    #print(table_rows)
    l = []
    title = {
        'A1': "Abituriyent ID raqami",
        'B1': "F.I.SH",
        'C1': "To'plagan ball",
        'D1': "Ta'lim tili",
        'E1': "Ta'lim shakli",
        'F1': "1-OTM",
        'G1': "2-OTM",
        'H1': "3-OTM",
        'I1': "4-OTM",
        'J1': "5-OTM"
    }
    for q, w in title.items():
        ws[q] = w
    columns = ['', 'A', 'B', '', '', 'C', 'D', 'E']
    for nr, tr in enumerate(table_rows[1:], 2):
        print(nr)
        nr = (i * 10 - 10) + nr
        td = tr.find_all('td')
        for nc, col in enumerate(td, 1):
            if nc in [1, 2, 5, 6, 7]:
                # print(columns[nc] + str(nr), col.text)
                ws[columns[nc] + str(nr)] = col.text
            elif nc == 8:
                shaxsiy_url = 'https://mandat.dtm.uz' + str(col.find('a')['href'])
                page = requests.get(shaxsiy_url, headers=headers)
                soup2 = BeautifulSoup(page.text, 'html.parser')
                soup3 = soup2.find_all('table', class_='table table-responsive table-bordered')[0].find_all('tr')
                columns1 = ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']
                for y, child in enumerate(soup3[2:]):
                    row = {}
                    gk = 0
                    child = child.find_all('td')
                    otm_fac = ''
                    for x, td in enumerate(child[1:3]):
                        try:
                            tx = td.text.replace('\n', '')
                            if x == 0:
                                otm_fac += tx
                            if x == 1:
                                otm_fac += ', ' + tx
                                if "Toshkent" in otm_fac and "sohalar" in otm_fac:
                                    cel = ws[columns1[y] + str(nr)]
                                    cel.fill = PatternFill("solid", fgColor="00FF0000")
                                ws[columns1[y] + str(nr)] = otm_fac
                                otm_fac = ''
                        except:
                            continue
        sleep(0.1)

wb.save('narxoz_iqtisod_kunduzgi.xlsx')

